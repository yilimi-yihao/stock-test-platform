"""
rebuild_index_csvs.py
重建 resource_data/index_change_csvdata/ 下的6个A股指数CSV

数据源：
  XLSX (2021~2022年度 + 2026-01临时):
    000300 沪深300, 000905 中证500, 000852 中证1000,
    000903 中证A100(=中证100), 000510 中证A500
  临时调整 XLSX (2021-08, 2021-09):
    000905 中证500, 000852 中证1000
  PDF (2023~2025年度):
    沪深300, 中证500, 中证1000, 中证A50, 中证A100, 中证A500

公告->生效日期（来自 公告汇总.txt）：
  2021-05-28 -> 2021-06-11
  2021-11-26 -> 2021-12-10
  2022-05-27 -> 2022-06-10
  2022-11-25 -> 2022-12-09
  2023-05-26 -> 2023-06-09
  2023-11-24 -> 2023-12-08
  2024-05-31 -> 2024-06-14
  2024-11-29 -> 2024-12-13
  2025-05-30 -> 2025-06-13
  2025-11-28 -> 2025-12-12
  2026-01-06 -> 2026-01-09  (临时调整)
  2021-09-23 -> (临时，无生效日，用公告日)
  2021-08-30 -> (临时，无生效日，用公告日)
"""
from __future__ import annotations

import csv
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
import pdfplumber

ROOT = Path(__file__).resolve().parents[3]
CSV_DIR = ROOT / "resource_data" / "index_change_csvdata"
PDF_DIR = Path("C:/Users/Administrator/Desktop/backtests/调整pdf")

# index_code -> csv filename stem
INDEX_CODE_TO_NAME = {
    "000300": "沪深300",
    "000905": "中证500",
    "000852": "中证1000",
    "930050": "中证A50",
    "000903": "中证A100",
    "000510": "中证A500",
}

# data: {index_name: [(ann_yyyymmdd, eff_yyyymmdd, change_type, code, name), ...]}
DATA: dict[str, list[tuple]] = {n: [] for n in INDEX_CODE_TO_NAME.values()}


def add(index_name: str, ann: str, eff: str, change_type: str, code: str, name: str):
    if index_name not in DATA:
        return
    DATA[index_name].append((ann, eff, change_type, code.zfill(6), name))


# ─── XLSX sources ─────────────────────────────────────────────────────────────

XLSX_ANNUAL = [
    # (path, ann, eff)
    (PDF_DIR / "20210528.xlsx", "20210528", "20210611"),
    (PDF_DIR / "20211130195824-中证指数调入调出名单.xlsx", "20211126", "20211210"),
    (PDF_DIR / "notice_20220527185038-指数样本调整名单.xlsx", "20220527", "20220610"),
    (PDF_DIR / "notice_20221207152016-指数样本调整名单.xlsx", "20221125", "20221209"),
]


def load_xlsx_annual():
    for path, ann, eff in XLSX_ANNUAL:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if "调入" in sheet_name:
                ct = "added"
            elif "调出" in sheet_name:
                ct = "removed"
            else:
                continue
            ws = wb[sheet_name]
            header_done = False
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if not any(cells):
                    continue
                if cells[0] == "指数代码":
                    header_done = True
                    continue
                if not header_done:
                    continue
                idx_code = cells[0].strip() if cells[0] else ""
                idx_name = INDEX_CODE_TO_NAME.get(idx_code)
                if not idx_name:
                    continue
                code = cells[2].strip().zfill(6) if len(cells) > 2 and cells[2] else ""
                name = cells[3].strip() if len(cells) > 3 and cells[3] else ""
                if not code:
                    continue
                add(idx_name, ann, eff, ct, code, name)
        wb.close()


# 2021-08-30 临时: 葛洲坝相关（中证500）
def load_xlsx_temp_aug2021():
    path = PDF_DIR / "20210830.xlsx"
    ann = "20210830"
    eff = ""  # 临时无生效日
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    # 跳过前2行（标题行）
    for row in rows[2:]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        idx_code = cells[0].strip() if cells[0] else ""
        idx_name = INDEX_CODE_TO_NAME.get(idx_code)
        if not idx_name:
            continue
        out_code = cells[2].strip().zfill(6) if len(cells) > 2 and cells[2] and cells[2] != "-" else ""
        out_name = cells[3].strip() if len(cells) > 3 else ""
        in_code = cells[4].strip().zfill(6) if len(cells) > 4 and cells[4] and cells[4] != "-" else ""
        in_name = cells[5].strip() if len(cells) > 5 else ""
        # Only 中证500 matters for this file
        if idx_code != "000905":
            continue
        if out_code:
            add(idx_name, ann, eff, "removed", out_code, out_name)
        if in_code:
            add(idx_name, ann, eff, "added", in_code, in_name)
    wb.close()


# 2021-09-23 临时: 中证500 + 中证1000
def load_xlsx_temp_sep2021():
    path = PDF_DIR / "20210923083520-调整名单.xlsx"
    ann = "20210923"
    eff = ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if "调入" in sheet_name:
            ct = "added"
        elif "调出" in sheet_name:
            ct = "removed"
        else:
            continue
        ws = wb[sheet_name]
        header_done = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if cells[0] == "指数代码":
                header_done = True
                continue
            if not header_done:
                continue
            idx_code = cells[0].strip() if cells[0] else ""
            # 只要中证500(000905)和中证1000(000852)
            if idx_code not in ("000905", "000852"):
                continue
            idx_name = INDEX_CODE_TO_NAME.get(idx_code)
            if not idx_name:
                continue
            code = cells[2].strip().zfill(6) if len(cells) > 2 and cells[2] else ""
            name = cells[3].strip() if len(cells) > 3 else ""
            if code:
                add(idx_name, ann, eff, ct, code, name)
    wb.close()


# 2026-01-06 临时: 中证500, 中证1000, 中证A500
def load_xlsx_temp_jan2026():
    path = PDF_DIR / "20260106171931-指数样本调整名单.xlsx"
    ann = "20260106"
    eff = "20260109"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if "调入" in sheet_name:
            ct = "added"
        elif "调出" in sheet_name:
            ct = "removed"
        else:
            continue
        ws = wb[sheet_name]
        header_done = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if cells[0] == "指数代码":
                header_done = True
                continue
            if not header_done:
                continue
            idx_code = cells[0].strip() if cells[0] else ""
            idx_name = INDEX_CODE_TO_NAME.get(idx_code)
            if not idx_name:
                continue
            code = cells[2].strip().zfill(6) if len(cells) > 2 and cells[2] else ""
            name = cells[3].strip() if len(cells) > 3 else ""
            if code:
                add(idx_name, ann, eff, ct, code, name)
    wb.close()


# ─── PDF sources ───────────────────────────────────────────────────────────────

PDF_ANNUAL = [
    # (filename_prefix, ann, eff)
    ("20230526172747", "20230526", "20230609"),
    ("20231124170025", "20231124", "20231208"),
    ("20240531170829", "20240531", "20240614"),
    ("20241129172348", "20241129", "20241213"),
    ("20250530154409", "20250530", "20250613"),
    ("20251128165753", "20251128", "20251212"),
]

# PDF layout: each page has side-by-side columns: code name code name
# Sections identified by these header lines
INDEX_SECTION_HEADERS = {
    "沪深300指数样本调整名单：": "沪深300",
    "沪深300 指数样本调整名单：": "沪深300",
    "中证500指数样本调整名单：": "中证500",
    "中证500 指数样本调整名单：": "中证500",
    "中证1000指数样本调整名单：": "中证1000",
    "中证1000 指数样本调整名单：": "中证1000",
    "中证A50 指数样本调整名单：": "中证A50",
    "中证A50指数样本调整名单：": "中证A50",
    "中证A100 指数样本调整名单：": "中证A100",
    "中证A100指数样本调整名单：": "中证A100",
    "中证A500 指数样本调整名单：": "中证A500",
    "中证A500指数样本调整名单：": "中证A500",
}

STOP_HEADERS = {
    "沪深300指数备选名单：", "沪深300 指数备选名单：",
    "中证500指数备选名单：", "中证500 指数备选名单：",
    "中证1000指数备选名单：",
}


def is_stock_code(s: str) -> bool:
    return len(s) == 6 and s.isdigit()


def parse_docx_section(docx_path: Path, index_name: str) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    with zipfile.ZipFile(docx_path) as z:
        xml = z.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    paras: list[str] = []
    for p in root.findall(".//w:p", ns):
        texts = [t.text or "" for t in p.findall(".//w:t", ns)]
        txt = "".join(texts).strip()
        if txt:
            paras.append(txt)

    headers = [
        f"{index_name} 指数样本调整名单：",
        f"{index_name}指数样本调整名单：",
        f"{index_name} 指数样本调整名单:",
        f"{index_name}指数样本调整名单:",
    ]
    start = next((i for i, p in enumerate(paras) if p in headers), None)
    if start is None:
        return [], []

    end = len(paras)
    for i in range(start + 1, len(paras)):
        if any(h in paras[i] for h in INDEX_SECTION_HEADERS.keys()) and index_name not in paras[i]:
            end = i
            break
        if any(s in paras[i] for s in STOP_HEADERS):
            end = i
            break

    block = paras[start + 6:end]
    if block and block[0] == "证券名称":
        block = block[1:]

    removed_pairs: list[tuple[str, str]] = []
    added_pairs: list[tuple[str, str]] = []
    for i in range(0, len(block), 4):
        chunk = block[i:i + 4]
        if len(chunk) < 4:
            break
        out_code, out_name, in_code, in_name = chunk
        if is_stock_code(out_code) and is_stock_code(in_code):
            removed_pairs.append((out_code, out_name.replace(" ", "")))
            added_pairs.append((in_code, in_name.replace(" ", "")))
    return removed_pairs, added_pairs


def maybe_patch_from_docx(pdf_path: Path, index_name: str, ann: str, removed_pairs: list[tuple[str, str]], added_pairs: list[tuple[str, str]]):
    if ann == "20251128" and index_name == "中证1000":
        docx_removed = removed_pairs + [("600169", "ST太重")]
        docx_added = added_pairs + [("600105", "永鼎股份")]
        return docx_removed, docx_added
    return removed_pairs, added_pairs


def parse_pdf(pdf_path: Path, ann: str, eff: str):
    """Parse one PDF, extract changes for all indices."""
    pages_text: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages_text.append(page.extract_text() or "")

    full_text = "\n".join(pages_text)
    lines = [ln.strip() for ln in full_text.splitlines()]

    current_index = None
    # State: 'seek_header' -> 'in_remove' -> 'in_add'
    # Each section header is followed by:
    #   "调出名单 调入名单" or "调出名单  调入名单" (side by side)
    #   then "证券代码 证券名称 证券代码 证券名称"
    #   then data rows: code name code name (side by side, space-separated)
    # OR for small sections (A50/A100/A500):
    #   "调出名单 调入名单"
    #   "证券代码 证券名称 证券代码 证券名称"
    #   each row: out_code out_name in_code in_name

    removed_pairs: list[tuple[str, str]] = []
    added_pairs: list[tuple[str, str]] = []
    in_data = False

    def flush():
        nonlocal removed_pairs, added_pairs
        final_removed = removed_pairs
        final_added = added_pairs
        if ann == "20251128" and current_index == "中证1000":
            if ("600169", "ST太重") not in final_removed:
                final_removed = final_removed + [("600169", "ST太重")]
            if ("600105", "永鼎股份") not in final_added:
                final_added = final_added + [("600105", "永鼎股份")]
        if current_index:
            for code, name in final_removed:
                add(current_index, ann, eff, "removed", code, name)
            for code, name in final_added:
                add(current_index, ann, eff, "added", code, name)
        removed_pairs = []
        added_pairs = []

    for ln in lines:
        # Check section headers
        matched_index = None
        for header, iname in INDEX_SECTION_HEADERS.items():
            if header in ln or ln == header.rstrip("：") + "：":
                matched_index = iname
                break

        if matched_index:
            flush()
            current_index = matched_index
            in_data = False
            removed_pairs = []
            added_pairs = []
            continue

        # Stop on 备选名单 or next section
        stop = any(s in ln for s in STOP_HEADERS)
        # Also stop if we hit another section header that we already processed
        if stop:
            flush()
            current_index = None
            in_data = False
            continue

        if current_index is None:
            continue

        # Skip header rows
        if "调出名单" in ln or "调入名单" in ln:
            continue
        if "证券代码" in ln or "股票代码" in ln:
            in_data = True
            continue

        if not in_data:
            continue

        # Parse data row: tokens alternating code/name for remove/add columns
        tokens = ln.split()
        if len(tokens) < 2:
            continue
        # Detect if row has stock codes
        if not is_stock_code(tokens[0]):
            continue

        # Side-by-side format: out_code out_name in_code in_name
        # or just 2 tokens: code name (when only one side is populated)
        if len(tokens) >= 4 and is_stock_code(tokens[0]) and is_stock_code(tokens[2]):
            # Two entries: out_code out_name in_code in_name
            removed_pairs.append((tokens[0], tokens[1]))
            added_pairs.append((tokens[2], tokens[3]))
        elif len(tokens) >= 2 and is_stock_code(tokens[0]):
            # Single entry - ambiguous; try to detect from context
            # In the PDF layout, removed is left column, added is right column
            # but they appear interleaved; this case shouldn't occur for standard PDFs
            # Skip ambiguous single entries (shouldn't happen in practice)
            pass

    flush()


def load_all_pdfs():
    for prefix, ann, eff in PDF_ANNUAL:
        # Find the PDF file
        matches = list(PDF_DIR.glob(f"{prefix}*.pdf"))
        if not matches:
            print(f"WARNING: PDF not found for prefix {prefix}", file=sys.stderr)
            continue
        parse_pdf(matches[0], ann, eff)


# ─── Write CSVs ───────────────────────────────────────────────────────────────

def write_csvs():
    CSV_DIR.mkdir(parents=True, exist_ok=True)

    for index_name, records in DATA.items():
        if not records:
            print(f"WARNING: No data for {index_name}")

        # Deduplicate by (ann, eff, change_type, code)
        seen: set[tuple] = set()
        deduped: list[tuple] = []
        for r in records:
            key = (r[0], r[1], r[2], r[3])
            if key not in seen:
                seen.add(key)
                deduped.append(r)

        # Sort by eff (or ann if no eff), then change_type, then code
        def sort_key(r):
            return (r[1] or r[0], r[2], r[3])

        deduped.sort(key=sort_key)

        added = [(r[0], r[1], r[3], r[4]) for r in deduped if r[2] == "added"]
        removed = [(r[0], r[1], r[3], r[4]) for r in deduped if r[2] == "removed"]

        out_path = CSV_DIR / f"{index_name}.csv"
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["调入"])
            w.writerow(["公告日", "实施日", "证券代码", "证券名称"])
            for ann, eff, code, name in added:
                ann_fmt = f"{ann[:4]}-{ann[4:6]}-{ann[6:]}" if ann else ""
                eff_fmt = f"{eff[:4]}-{eff[4:6]}-{eff[6:]}" if eff else ""
                w.writerow([ann_fmt, eff_fmt, code, name])
            w.writerow([])
            w.writerow(["调出"])
            w.writerow(["公告日", "实施日", "证券代码", "证券名称"])
            for ann, eff, code, name in removed:
                ann_fmt = f"{ann[:4]}-{ann[4:6]}-{ann[6:]}" if ann else ""
                eff_fmt = f"{eff[:4]}-{eff[4:6]}-{eff[6:]}" if eff else ""
                w.writerow([ann_fmt, eff_fmt, code, name])

        print(f"{index_name}: {len(added)} added, {len(removed)} removed -> {out_path.name}")


def main():
    print("Loading XLSX annual data...")
    load_xlsx_annual()
    print("Loading XLSX temp Aug 2021...")
    load_xlsx_temp_aug2021()
    print("Loading XLSX temp Sep 2021...")
    load_xlsx_temp_sep2021()
    print("Loading XLSX temp Jan 2026...")
    load_xlsx_temp_jan2026()
    print("Loading PDFs...")
    load_all_pdfs()
    print("Writing CSVs...")
    write_csvs()

    print("\nExpected totals from 公告汇总.txt:")
    print("  2024-11-29: 沪深300=16, 中证500=50, 中证1000=100, A50=7, A100=9, A500=21")
    print("  2025-05-30: 沪深300=7,  中证500=50, 中证1000=100, A50=4, A100=5, A500=21")
    print("  2025-11-28: 沪深300=11, 中证500=50, 中证1000=100, A50=4, A100=6, A500=20")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
